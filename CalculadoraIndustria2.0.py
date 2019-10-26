import openpyxl as xl
import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename

class Dias:
    def __init__(self, date, row, entrada1, saida1, entrada2, saida2, ehFalta = False):
        self.date = date #Dias (Dias, mes ano)
        self.row = row
        self.entrada1 = entrada1 #Horário que entrou
        self.saida1 = saida1 #Horário que saiu para o almoço, se for meio período, horário que saiu 
        self.entrada2 = entrada2 #Horário que voltou para o almoço
        self.saida2 = saida2 #Horário que foi embora
        #Regra, horário de entrada ideal, tempo de almoço, horário de saída ideal, 
        #à partir de que horário recebe adicional noturno
        self.ehFalta= ehFalta

   
class Funcionario:
    def __init__ (self, name):
        self.name = name
        self.Dias = {}
        self.Regras = {}
        self.HoraExtraCem = 0 #Total referente a hora extra 100% do salário
        self.TotalFaltas = 0 #Total de dias que o funcionario faltou
        self.TotalAtraso = 0 #Tempo total que ele chegou atrasado
        self.TotalAdianto = 0 #Tempo total que le chegou adiantado
        self.TotalDeAdicional = 0 #Total de Adicional noturno
        pass
    pass
    
class Rules: #New Rules
    def __init__ (self, date, entradaIdeal, saidaIdeal, TempoDeAlmoço,TotalDeAdicional, ehFeriado=False, meioperiodo=False, ehFolga=False):    
        self.date = date
        self.entradaIdeal=entradaIdeal
        self.saidaIdeal=saidaIdeal
        self.TempoDeAlmoço=TempoDeAlmoço
        self.TotalDeAdicional=TotalDeAdicional
        self.ehFeriado = ehFeriado
        self.meioperiodo = meioperiodo
        self.ehFolga = ehFolga
        pass
   
    pass
#Converte um numero inteiro para um formato horário    
def ConvertToTime (inteiro):
    if inteiro == None:
        return 0
    h = int(inteiro/100)
    m = inteiro%100
    time = h/24 + m/1440
    return time
#Converte um formato horário para um número inteiro
def ConvertToNumber (tempo):
    if tempo == None:
        return 0
    tempoH = tempo*24
    h = int(tempoH)
    m = round((tempoH-h)*60)
    return h*100+m
def FullfillFeriados (Feriados):
    FindedFeriados = False
    Feriados_row = 0
    Feriados_column = 0
    for column in range (1,Feriados.max_column):
        for row in range(1,Feriados.max_row):
            if Feriados.cell(row=row,column=column).value=="FERIADOS":               
                FindedFeriados = True
                Feriados_row = row
                Feriados_column = column
    if (FindedFeriados == False):
        return
    row = Feriados_row+1
    ContadorDias = 0
    while Feriados.cell(row = row, column = Feriados_column).value != None:
        feriados.append(Feriados.cell(row=row,column=Feriados_column).value)
        row += 1
        ContadorDias +=1
 
def preenchendoDados (Func, Regra):
    alldomingos = True
    #Criando Variaveis
    Data_row = 0  #linha onde começam as datas, começa para valer em Data_row+1
    Data_column = 0 #Coluna onde começam as datas, começa para valer em Data_column+1
    #define o nome da tabela que é o nome do funcionario, 4 primeiras letras sao Func
    funcionario = Funcionario(Func.title[4:])
    print (funcionario.name)    
    #Procura o inicio da tabela pelo nome Data###############################
    FindedData1 = False
    for column in range (1,Func.max_column):
        for row in range(1,Func.max_row):
            if Func.cell(row=row,column=column).value=="Data":
                Data_row = row
                Data_column = column
                FindedData1 = True
            if (FindedData1):
                break
        if (FindedData1):
            break
    if (Data_row==0 or Data_column==0):
        print ("Error 'Data' não encontrada, por favor recheque sua planilha")
        exit
    ##############Achamos onde esta a Data#############################################
    
    #preenche os valores da tabela
    row = Data_row+1 #Eu disse que o real era +1
    ContadorDias = 0 #ContaQuantosDiasTem    
    
    ##############Preenche os valores da tabela########################################
    row = Data_row+1 #Eu disse que o real era +1
    ContadorDias = 0 #ContaQuantosDiasTem
    #Preenchendo informação do dia e da regra do dia
    #Dia e Dia da semana Ideal
    diaI = None
    semana = {0:None, 1:None,2:None,3:None,4:None,5:None,6:None}
    while Func.cell(row = row, column = Data_column).value != None:
        NovoOuCopia=True
        #Informação do Dia
        Data = (Func.cell(row=row,column=Data_column).value)
        entrada1 = (Func.cell(row = row, column = Data_column+2).value)
        saida1 = (Func.cell(row = row, column = Data_column+3).value)
        entrada2 = (Func.cell(row = row, column = Data_column+4).value)
        saida2 = (Func.cell(row = row, column = Data_column+5).value)  
        #Informação da Regra do Dia
        #ver se é feriado
        if (Data in feriados):
            if(Regra.cell(row=row, column=Data_column+2).value == None):
                Regra.cell(row = row, column = Data_column+2, value="FERIADO")
        entradaIdeal = (Regra.cell(row = row, column = Data_column+2).value)
        saidaIdeal = (Regra.cell(row = row, column = Data_column+3).value)
        TempoDeAlmoço = (Regra.cell(row = row, column = Data_column+4).value)
        horarioDoAdicional = (Regra.cell(row = row, column = Data_column+5).value)
        meioperiodo = (Regra.cell(row=row,column=Data_column+6).value)
        #Criando objetos dia e regra do tipo Dia e Regras        
        if(entradaIdeal is None):
            NovoOuCopia=False
            if(semana[Data.weekday()] is None):
                if(diaI is None):
                    exit('tabela de regras incompleta')
                entradaIdeal=diaI.entradaIdeal
                saidaIdeal=diaI.saidaIdeal
                TempoDeAlmoço=diaI.TempoDeAlmoço
                horarioDoAdicional=diaI.TotalDeAdicional
                if(diaI.meioperiodo):
                    meioperiodo=True
                else:
                    meioperiodo=None
            else:
                entradaIdeal=semana[Data.weekday()].entradaIdeal
                saidaIdeal=semana[Data.weekday()].saidaIdeal
                TempoDeAlmoço=semana[Data.weekday()].TempoDeAlmoço
                horarioDoAdicional=semana[Data.weekday()].TotalDeAdicional
                if(semana[Data.weekday()].meioperiodo):
                    meioperiodo=True
                else:
                    meioperiodo=None
        
        #Verificaremos se regra é feriado, férias, folga ou atestado        
        if (entradaIdeal == 'FERIADO'):
            dia = Dias(Data, row, entrada1, saida1, entrada2, saida2) 
            regra = Rules(Data, entradaIdeal, saidaIdeal, TempoDeAlmoço, horarioDoAdicional, True)
        elif (entradaIdeal == 'FOLGA' or entradaIdeal=='ATESTADO' or entradaIdeal=='FÉRIAS' or entrada1 == 'ATESTADO' or entrada1=='FOLGA' or entrada1 == 'FÉRIAS'):        
            regra = Rules(Data, entradaIdeal, saidaIdeal, TempoDeAlmoço, horarioDoAdicional, ehFolga=True)
            dia = Dias(Data, row, entrada1, saida1, entrada2, saida2)
            if (Data.weekday() == 6):
                alldomingos = False
        #Verificaremos se dia é falta
        elif (entrada1 == 'FALTA'):
            dia = Dias(Data, row, entrada1, saida1, entrada2, saida2, True)
            regra = Rules(Data, entradaIdeal, saidaIdeal, TempoDeAlmoço, horarioDoAdicional)
            if (Data.weekday() == 6):
                alldomingos = False
        elif(meioperiodo is not None):
            dia = Dias(Data, row, entrada1, saida1, entrada2, saida2)
            regra = Rules(Data, entradaIdeal, saidaIdeal, TempoDeAlmoço, horarioDoAdicional,meioperiodo=True)    
        else:
            dia = Dias(Data, row, entrada1, saida1, entrada2, saida2)
            regra = Rules(Data, entradaIdeal, saidaIdeal, TempoDeAlmoço, horarioDoAdicional)
        #Adicionando dois dicionarios, um de dia e um de regra
        funcionario.Dias.update({Data: dia})
        funcionario.Regras.update({Data: regra})
        #Criando Dia e semana ideais
        if(diaI is None and NovoOuCopia):
            if(regra.ehFeriado is False):
                diaI=regra
        if(semana[Data.weekday()] is None and NovoOuCopia):
            if(regra.ehFeriado is False and regra.ehFolga is False):
                semana.update({Data.weekday():regra})
        row += 1
        ContadorDias +=1

    ##############Calculo da hora extra e atrasos etc ###################################
    for i in funcionario.Dias:
    #Caso seja feriado
        if (funcionario.Regras[i].ehFeriado):
            if isinstance(funcionario.Dias[i].entrada1, int) or isinstance(funcionario.Dias[i].entrada2, int):
                temp = -ConvertToTime(funcionario.Dias[i].entrada1)+ConvertToTime(funcionario.Dias[i].saida1) - ConvertToTime(funcionario.Dias[i].entrada2) + ConvertToTime(funcionario.Dias[i].saida2)
                funcionario.HoraExtraCem+=temp
    #Caso seja folga
        elif (funcionario.Regras[i].ehFolga):
            if isinstance(funcionario.Dias[i].entrada1, int) or isinstance(funcionario.Dias[i].entrada2, int):
                    temp = -ConvertToTime(funcionario.Dias[i].entrada1)+ConvertToTime(funcionario.Dias[i].saida1) - ConvertToTime(funcionario.Dias[i].entrada2) + ConvertToTime(funcionario.Dias[i].saida2)
                    if(dia.date.weekday()==6):
                        funcionario.HoraExtraCem+=temp
                    else:
                        funcionario.TotalAdianto+=temp
        #Caso tenha faltado
        elif (funcionario.Dias[i].ehFalta):
            funcionario.TotalFaltas+=1
    #Caso contrário
        else:
            regra = funcionario.Regras[i]
            dia = funcionario.Dias[i]
            #verificando meio-periodo
            if (dia.entrada1 is None and dia.saida1 is None):
                dia.entrada1=dia.entrada2
                dia.saida1 = dia.saida2
            elif (dia.entrada2 is None and dia.saida2 is None):
                #Entrada
                temp = ConvertToTime(dia.entrada1)-ConvertToTime(regra.entradaIdeal)
                if (temp>5/1440):
                    funcionario.TotalAtraso+=temp
                elif (temp<-5/1440):
                    if(dia.date.weekday()==6):
                        funcionario.HoraExtraCem-=temp
                    else:
                        funcionario.TotalAdianto-=temp
                else:
                    temp = 0
                Func.cell(row=dia.row, column=Data_column+6, value = ConvertToNumber(-temp))
                #Saida
                if (dia.saida1>500):
                    temp = ConvertToTime(dia.saida1)-ConvertToTime(regra.saidaIdeal)
                else:
                    temp = 1.00+ ConvertToTime(dia.saida1-regra.saidaIdeal)
                if (temp>5/1440):
                    if(dia.date.weekday()==6):
                        funcionario.HoraExtraCem+=temp
                    else:
                        funcionario.TotalAdianto+=temp
                elif (temp<-5/1440):
                    funcionario.TotalAtraso-=temp
                else:
                    temp = 0
                Func.cell(row=dia.row, column=Data_column+8, value = ConvertToNumber(temp))
                #Adicional Noturno
                if (regra.TotalDeAdicional is not None):
                    if (dia.saida1>500):
                        temp = ConvertToTime(dia.saida1)-ConvertToTime(regra.TotalDeAdicional)
                    else:
                        temp = ConvertToTime(dia.saida1)-ConvertToTime(regra.TotalDeAdicional)+1.0
                    if (temp<0):
                        temp=0
                    funcionario.TotalDeAdicional+=temp
                    Func.cell(row=dia.row, column=Data_column+9, value = ConvertToNumber(temp))
                            
            #Caso não seja meio-período
            else:
                #Entrada
                temp = ConvertToTime(dia.entrada1)-ConvertToTime(regra.entradaIdeal)
                if (temp>5/1440):
                    funcionario.TotalAtraso+=temp
                elif (temp<-5/1440):
                    if(dia.date.weekday()==6):
                        funcionario.HoraExtraCem-=temp
                    else:
                        funcionario.TotalAdianto-=temp
                else:
                    temp = 0
                Func.cell(row=dia.row, column=Data_column+6, value = ConvertToNumber(-temp))
                #Caso Meio periodo seja positivo vamos descontar o horário de almoço do tempo extra
                if(regra.meioperiodo):
                    #Saida
                    if (dia.saida2>500):
                        temp = ConvertToTime(dia.saida2)-ConvertToTime(regra.saidaIdeal)-(-ConvertToTime(dia.saida1)+ConvertToTime(dia.entrada2))
                    else:
                        temp = 1.00+ ConvertToTime(dia.saida2)-ConvertToTime(regra.saidaIdeal)
                    if (temp>5/1440):
                        if(dia.date.weekday()==6):
                            funcionario.HoraExtraCem+=temp
                        else:
                            funcionario.TotalAdianto+=temp
                    elif (temp<-5/1440):
                        funcionario.TotalAtraso-=temp
                    else:
                        temp = 0
                    Func.cell(row=dia.row, column=Data_column+8, value = ConvertToNumber(temp))  
                #Caso Contrário
                else:
                    #Saida
                    if (dia.saida2>500):
                        temp = ConvertToTime(dia.saida2)-ConvertToTime(regra.saidaIdeal)
                    else:
                        temp = 1.00+ ConvertToTime(dia.saida2)-ConvertToTime(regra.saidaIdeal)
                    if (temp>5/1440):
                        if(dia.date.weekday()==6):
                            funcionario.HoraExtraCem+=temp
                        else:
                            funcionario.TotalAdianto+=temp
                    elif (temp<-5/1440):
                        funcionario.TotalAtraso-=temp
                    else:
                        temp = 0
                    Func.cell(row=dia.row, column=Data_column+8, value = ConvertToNumber(temp))  
                    #Almoço
                    temp = ConvertToTime(dia.entrada2)-ConvertToTime(dia.saida1)-ConvertToTime(regra.TempoDeAlmoço)
                    if (temp>5/1440):
                        funcionario.TotalAtraso+=temp
                    elif (temp<-5/1440):
                        if(dia.date.weekday()==6):
                            funcionario.HoraExtraCem-=temp
                        else:
                            funcionario.TotalAdianto-=temp
                    else:
                        temp = 0
                    Func.cell(row=dia.row, column=Data_column+7, value = ConvertToNumber(-temp))
                #Adicional Noturno
                if (regra.TotalDeAdicional is not None):
                    if (dia.saida2>500):
                        temp = ConvertToTime(dia.saida2)-ConvertToTime(regra.TotalDeAdicional)
                    else:
                        temp = ConvertToTime(dia.saida2)-ConvertToTime(regra.TotalDeAdicional)+1.0
                    if (temp<0):
                        temp=0
                    funcionario.TotalDeAdicional+=temp
                    Func.cell(row=dia.row, column=Data_column+9, value = ConvertToNumber(temp))
    #Printar Resultados
    Func.cell(row =46, column = 5, value = funcionario.name)
    Func.cell(row=46, column = 6, value = "Extra 70%")
    Func.cell(row=46, column = 7, value = "Atrasos")
    Func.cell(row=46, column = 8, value = "Adc. Noturno")
    Func.cell(row=46, column = 10, value = "Faltas")
    Func.cell(row=47, column = 5, value = "Horário normal")
    Func.cell(row=48, column = 5, value = "Horário em h")
    Func.cell(row=46, column =9, value = "Extra 100%")
    Func.cell(row=46, column =11, value = "Domingo Mês")
    Func.cell(row=47, column = 6, value = ConvertToNumber(funcionario.TotalAdianto))
    Func.cell(row=47, column = 7, value = ConvertToNumber(funcionario.TotalAtraso))
    Func.cell(row=47, column = 8, value = ConvertToNumber(funcionario.TotalDeAdicional))
    Func.cell(row=47, column = 10, value = funcionario.TotalFaltas)
    Func.cell(row=47, column =9, value = ConvertToNumber(funcionario.HoraExtraCem))
    if (alldomingos==True):Func.cell(row=47, column =11, value = "Paga")
    else: Func.cell(row=47, column =11, value = "Não Paga")                                   
    Func.cell(row=48, column = 6, value = "Tempo em horas")
    Func.cell(row=48, column = 6, value = (funcionario.TotalAdianto)*24)
    Func.cell(row=48, column = 7, value = (funcionario.TotalAtraso)*24)
    Func.cell(row=48, column = 8, value = (funcionario.TotalDeAdicional)*24*60/52.5)
    Func.cell(row=48, column =9, value = (funcionario.HoraExtraCem)*24)
    
def main():
    print ("Calculadora de tempo trabalhado - developed by Melo")
    print ("Digite o nome do arquivo")
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    nome = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    print(nome)
    wb = xl.load_workbook(filename = nome, keep_vba = True, read_only = False)
    ws = {}
    FullfillFeriados(wb["CAPA"])
    for Func in wb.sheetnames:
        if Func[:4] == "Func":
            ws.update({Func:"Regra"+Func[4:]})
    print("nos encontramos os seguintes funcionarios: ", ws.items())
    print ("e as seguintes regras", ws.values())
    print ("cerifique-se que cada funcionario tem uma regra")
    yeswecan = input("podemos prosseguir:(Y/n) ")
    if yeswecan == 'n':
        exit ("operação cancelada")
    for Func in ws:
        preenchendoDados(wb[Func], wb[ws[Func]])
    print("Salvando o arquivo, por favor não feche o programa\n*sujeito a corromper arquivo*")
    wb.save(nome)
    print ("Operation Sucefull ocurred")
    
    

feriados = []
if __name__=="__main__":
    main()    
